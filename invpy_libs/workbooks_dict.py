#! python3
# Example to get all values based on column index
import openpyxl


# from openpyxl.cell import get_column_letter,column_index_from_string

def excel_get_dict(workbook, sheet, column_ref):
    '''

    :param workbook:   .xlsx file to read
    :param sheet:      sheet name on the workbook
    :param column_ref: column to use as main key for the dict.
    :return: dictionary with the data
    '''
    # Open the spreadsheet and get the latest dues status.
    wb = openpyxl.load_workbook(workbook)
    sheet = wb.get_sheet_by_name(sheet)
    # Define columns index by Name with value in row 1
    # Example: {'status': 'C', 'Notes': 'K'}
    columns = {}
    # for 1 to last column
    for i in range(1, sheet.get_highest_column()):
        # set dict columns.setdefault('value', 'column')
        columns.setdefault(sheet.cell(row=1, column=i).value, sheet.cell(row=1, column=i).column)
    # Define sheet dictionary.
    '''
    Create list of servers, example:
     'HName': {'IP': '192.168.69.5',
                   'OS': 'Vmware',
                   'description': 'Vmware esx server',
                   'status': 'active'}}
    '''

    datadict = {}
    # Read each row of the sheet.
    for row in range(2, sheet.get_highest_row() + 1):  # Skip the first row
        # read each column of each row.
        # k is key value, like: Name
        # v is value of the key, the column index like A
        for k, v in sorted(columns.items()):
            # Set dictionary for each server in nested dict
            # key server
            # Nested key names and values based on each column key Name
            # like datadict.setdefault('servername', {})['IP'] = sheet['A5'].value
            datadict.setdefault(sheet[columns.get(column_ref) + str(row)].value, {})[k] = sheet[v + str(row)].value
    return datadict

    # Example:
    # file = 'servers.xlsx'
    # servers = excel_get_dict(workbook, sheet, column_ref)
    # servers = excel_get_dict(file, 'Servers','Name')
